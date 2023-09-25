import openpyxl
import numpy as np


def get_roc_array(label0, label1):
    threpool = label0 + label1
    threpool.sort()
    label0 = np.array(label0)
    label1 = np.array(label1)
    tprlist = []
    fprlist = []
    max_th = 0
    max_tpr_fpr = 0
    TPout = 0
    FPout = 0
    TNout = 0
    FNout = 0
    sensout = 0
    accuout = 0
    f1scoreout = 0
    specout = 0
    ppvout = 0
    npvout = 0
    for threshold in threpool:
        TP = np.sum(label1 >= threshold)
        FP = np.sum(label0 >= threshold)
        TN = np.sum(label0 < threshold)
        FN = np.sum(label1 < threshold)

        spec = TN / (TN + FP)
        sens = TP / (TP + FN)

        TPR = sens
        FPR = 1 - spec
        tprlist.append(TPR)
        fprlist.append(FPR)

        if (TPR - FPR) > max_tpr_fpr:
            max_th = threshold
            max_tpr_fpr = TPR - FPR
            sensout = sens
            accuout = (TP + TN) / (TP + FN + FP + TN)
            f1scoreout = 2 * TP / (2 * TP + FP + FN)
            specout = spec
            ppvout = TP / (TP + FP)
            npvout = TN / (TN + FN)
            TPout = TP
            TNout = TN
            FPout = FP
            FNout = FN

    print('sensitivity: {}\naccuarcy: {}\n'
          'f1-score: {}\nspecificicy: {}\n'
          'ppv: {}\nnpv: {}\nthreshold: {}'.format(sensout, accuout, f1scoreout, specout, ppvout, npvout, max_th))
    print('{}  {}\n{}  {}'.format(TPout, FPout, FNout, TNout))
    return max_th


if __name__ == '__main__':

    excel_file = '../ann_branch/ann_predict_ln.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    label0 = []
    label1 = []
    for row in range(2, sheet.max_row + 1):
        gt = sheet.cell(row=row, column=3).value
        pred = sheet.cell(row=row, column=2).value
        if gt == 1:
            label1.append(pred)
        else:
            label0.append(pred)

    max_th = get_roc_array(label0, label1)
    print(excel_file, max_th)
