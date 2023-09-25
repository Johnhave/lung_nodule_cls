# --coding:utf-8--
import os
import openpyxl


def get_data_from_excel(root_dir):
    data_xlsx_dir = r'E:\lkf\课题交流文件\LVI+VPI+LN_data(non enhance CT)20230410（2）.xlsx'
    wb = openpyxl.load_workbook(data_xlsx_dir)
    sheet = wb.active
    patient = []

    ct_time = []
    image_file_all = []

    # Y_dict = {31: '脉管侵犯（0无，1有，x报告）',
    #           29: '脏层胸膜侵犯VPI（0无，1有）',
    #           26: '淋巴结转移（0无；1有）'}
    txt_file = open('dataset_all.txt', 'w')
    for row in range(2, 1898):
        lab12 = str(sheet.cell(row=row, column=12).value)
        if lab12 == '1':
            lab12 = 1
        else:
            lab12 = 0

        lab13 = str(sheet.cell(row=row, column=13).value)
        if lab13 == '1':
            lab13 = 1
        else:
            lab13 = 0

        lab14 = str(sheet.cell(row=row, column=14).value)
        if lab14 == '1':
            lab14 = 1
        else:
            lab14 = 0

        pt = '{:010d}'.format(sheet.cell(row = row, column = 3).value)
        nd = '①'


        time_tmp = str(sheet.cell(row=row, column=22).value).split(' ')[0].replace('-', '')
        img_file = os.path.join(root_dir, '{}_{}_lung_{}.npy'.format(pt, time_tmp, nd))
        if not os.path.exists(img_file):
            img_file = os.path.join(root_dir, '{}_{}_med_{}.npy'.format(pt, time_tmp, nd))
            if not os.path.exists(img_file):
                print(pt, time_tmp, nd)
                continue

        patient.append(pt)
        ct_time.append(time_tmp)
        txt_file.write('{},{},{},{}\n'.format(img_file, lab12, lab13, lab14))
        image_file_all.append(img_file)
    txt_file.close()

if __name__ == '__main__':
    get_data_from_excel(r'E:\lkf\data_pool_sysucc_gy')